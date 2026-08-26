import fcntl
import threading
from contextlib import contextmanager
from pathlib import Path
from typing import List, Dict, Any, Optional, Set, Tuple
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.config import settings

_THREAD_LOCK = threading.Lock()


@contextmanager
def _excel_file_lock(lock_path: Path):
    """Процессный + файловый лок, чтобы веб, бот и автопилот не портили xlsx одновременно."""
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    with _THREAD_LOCK:
        with open(lock_path, "a+", encoding="utf-8") as lock_file:
            fcntl.flock(lock_file.fileno(), fcntl.LOCK_EX)
            try:
                yield
            finally:
                fcntl.flock(lock_file.fileno(), fcntl.LOCK_UN)


class ExcelStorage:
    """Управление локальной базой вакансий в формате Excel (.xlsx)."""

    COLUMNS = [
        "id",
        "status",
        "match_score",
        "title",
        "employer",
        "city",
        "salary_str",
        "skills",
        "url",
        "published_at",
        "cover_letter",
        "notes",
        "description",
        "salary_from",
        "salary_to",
        "currency",
    ]

    COLUMN_NAMES_RU = {
        "id": "ID Вакансии",
        "status": "Статус",
        "match_score": "Score (%)",
        "title": "Название вакансии",
        "employer": "Компания",
        "city": "Город",
        "salary_str": "Зарплата",
        "skills": "Ключевые навыки",
        "url": "Ссылка",
        "published_at": "Дата публикации",
        "cover_letter": "Сопроводительное письмо",
        "notes": "Заметки / Резюме анализа",
        "description": "Полное описание",
        "salary_from": "ЗП от",
        "salary_to": "ЗП до",
        "currency": "Валюта",
    }

    def __init__(self, file_path: Path = settings.EXCEL_PATH):
        self.file_path = file_path
        self.lock_path = file_path.parent / ".excel.lock"

    def _id_col(self, df: pd.DataFrame) -> str:
        return "ID Вакансии" if "ID Вакансии" in df.columns else "id"

    def get_existing_ids(self) -> Set[str]:
        """Возвращает множество уже сохраненных ID вакансий, чтобы избежать дубликатов."""
        df = self.load_all()
        if df.empty:
            return set()
        col = self._id_col(df)
        if col not in df.columns:
            return set()
        return set(df[col].dropna().astype(str).tolist())

    def load_all(self) -> pd.DataFrame:
        """Загрузка всех сохраненных вакансий в DataFrame."""
        with _excel_file_lock(self.lock_path):
            return self._load_unlocked()

    def _load_unlocked(self) -> pd.DataFrame:
        if not self.file_path.exists():
            return pd.DataFrame(columns=list(self.COLUMN_NAMES_RU.values()))
        return pd.read_excel(self.file_path, dtype=str)

    def clear_all(self) -> bool:
        """Полная очистка локальной базы вакансий (создает чистый Excel-файл с заголовками)."""
        empty_df = pd.DataFrame(columns=list(self.COLUMN_NAMES_RU.values()))
        with _excel_file_lock(self.lock_path):
            self._write_styled_excel(empty_df)
        return True

    def save_or_update_vacancies(self, vacancies: List[Dict[str, Any]]) -> Tuple[int, int]:
        """
        Сохраняет новые вакансии и обновляет данные (описание, скор, навыки) для уже существующих.
        Возвращает (added_count, updated_count).
        """
        if not vacancies:
            return 0, 0

        with _excel_file_lock(self.lock_path):
            existing_df = self._load_unlocked()
            id_col = self._id_col(existing_df)

            existing_ids = set()
            if not existing_df.empty and id_col in existing_df.columns:
                existing_ids = set(existing_df[id_col].dropna().astype(str).tolist())

            new_items = []
            updated_count = 0

            for v in vacancies:
                v_id = str(v.get("id"))
                if v_id in existing_ids:
                    mask = existing_df[id_col].astype(str) == v_id
                    if mask.any():
                        for key, val in v.items():
                            ru_col = self.COLUMN_NAMES_RU.get(key, key)
                            if ru_col in existing_df.columns and val is not None and str(val).strip():
                                existing_df.loc[mask, ru_col] = str(val)
                        updated_count += 1
                else:
                    new_items.append(v)

            if new_items:
                new_df = pd.DataFrame(new_items)
                new_df = new_df.rename(columns=self.COLUMN_NAMES_RU)
                if not existing_df.empty:
                    combined_df = pd.concat([existing_df, new_df], ignore_index=True)
                else:
                    combined_df = new_df
            else:
                combined_df = existing_df

            self._write_styled_excel(combined_df)
            return len(new_items), updated_count

    def save_new_vacancies(self, vacancies: List[Dict[str, Any]]) -> int:
        """
        Добавляет новые вакансии и обновляет существующие.
        Возвращает число новых записей (если новых нет — число обновлённых).
        """
        added, updated = self.save_or_update_vacancies(vacancies)
        return added if added > 0 else updated

    def update_status(
        self,
        vacancy_id: str,
        status: Optional[str] = None,
        match_score: Any = None,
        cover_letter: str = "",
        notes: str = "",
    ) -> bool:
        """Обновление статуса, скора или сопроводительного письма для конкретной вакансии."""
        return self.update_rows(
            [
                {
                    "id": vacancy_id,
                    "status": status,
                    "match_score": match_score,
                    "cover_letter": cover_letter,
                    "notes": notes,
                }
            ]
        ) > 0

    def update_rows(self, updates: List[Dict[str, Any]]) -> int:
        """Пакетное обновление строк Excel одним проходом записи."""
        if not updates:
            return 0

        with _excel_file_lock(self.lock_path):
            if not self.file_path.exists():
                return 0
            df = self._load_unlocked()
            id_col = self._id_col(df)
            changed = 0

            for item in updates:
                vacancy_id = str(item.get("id") or "")
                if not vacancy_id:
                    continue
                mask = df[id_col].astype(str) == vacancy_id
                if not mask.any():
                    continue

                status = item.get("status")
                match_score = item.get("match_score")
                cover_letter = item.get("cover_letter", "")
                notes = item.get("notes", "")

                if status and "Статус" in df.columns:
                    df.loc[mask, "Статус"] = status
                elif status and "status" in df.columns:
                    df.loc[mask, "status"] = status
                if match_score is not None and "Score (%)" in df.columns:
                    df.loc[mask, "Score (%)"] = str(match_score)
                if cover_letter and "Сопроводительное письмо" in df.columns:
                    df.loc[mask, "Сопроводительное письмо"] = cover_letter
                if notes and "Заметки / Резюме анализа" in df.columns:
                    df.loc[mask, "Заметки / Резюме анализа"] = notes
                changed += 1

            if changed:
                self._write_styled_excel(df)
            return changed

    def _write_styled_excel(self, df: pd.DataFrame):
        """Красивое сохранение с авто-шириной столбцов и стилизацией шапки."""
        with pd.ExcelWriter(self.file_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Вакансии")
            worksheet = writer.sheets["Вакансии"]

            # Цвета и стили
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            border = Border(
                left=Side(style="thin", color="E0E0E0"),
                right=Side(style="thin", color="E0E0E0"),
                top=Side(style="thin", color="E0E0E0"),
                bottom=Side(style="thin", color="E0E0E0"),
            )

            # Форматирование шапки
            for col_idx, cell in enumerate(worksheet[1], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Настройка ширины колонок и выравнивания ячеек
            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                col_name = str(col[0].value)

                for cell in col:
                    cell.border = border
                    val_str = str(cell.value or "")
                    if cell.row > 1:
                        cell.font = Font(name="Arial", size=10)
                        # Выравнивание по левому краю для текста, по центру для статусов
                        if col_name in ["ID Вакансии", "Статус", "Score (%)", "Дата публикации", "Город"]:
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")

                    max_len = max(max_len, len(val_str.split("\n")[0]))

                # Задаем адаптивную ширину колонок
                if col_name in ["Название вакансии", "Компания"]:
                    worksheet.column_dimensions[col_letter].width = 30
                elif col_name in ["Зарплата", "Ключевые навыки"]:
                    worksheet.column_dimensions[col_letter].width = 25
                elif col_name in ["Сопроводительное письмо", "Заметки / Резюме анализа", "Полное описание"]:
                    worksheet.column_dimensions[col_letter].width = 40
                elif col_name == "Ссылка":
                    worksheet.column_dimensions[col_letter].width = 28
                else:
                    worksheet.column_dimensions[col_letter].width = max(max_len + 3, 14)
